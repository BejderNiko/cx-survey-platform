"""Round-trip and safety tests for the export/import adapters."""
from __future__ import annotations

import io
import json

import pandas as pd
import pytest

from ok_analytics.contracts import DatasetPayload, VariableMeta
from ok_analytics.exports import export_csv, export_json, export_sav, export_xlsx, import_sav


@pytest.fixture()
def dataset() -> DatasetPayload:
    return DatasetPayload(
        variables=[
            VariableMeta(name="respondent_key", label="Respondent", var_type="string", measure="nominal"),
            VariableMeta(name="nps_score", label="NPS 0-10", var_type="numeric", measure="scale",
                         value_labels={"10": "Extremely likely"}),
            VariableMeta(name="main_reason", label="Main reason", var_type="string", measure="nominal",
                         value_labels={"price": "Pris", "service": "Service"}),
        ],
        rows=[
            {"respondent_key": "r_1", "nps_score": 10, "main_reason": "price"},
            {"respondent_key": "r_2", "nps_score": 3, "main_reason": "service"},
            {"respondent_key": "r_3", "nps_score": None, "main_reason": None},
            {"respondent_key": "=SUM(A1:A9)", "nps_score": 7, "main_reason": "price"},
        ],
    )


def test_csv_roundtrip_and_formula_injection(dataset):
    raw = export_csv(dataset).decode("utf-8-sig")
    assert "'=SUM(A1:A9)" in raw, "formula cells must be neutralized"
    df = pd.read_csv(io.StringIO(raw))
    assert list(df.columns) == ["respondent_key", "nps_score", "main_reason"]
    assert len(df) == 4
    assert df.loc[0, "nps_score"] == 10


def test_json_roundtrip_preserves_metadata(dataset):
    payload = json.loads(export_json(dataset))
    assert payload["variables"][1]["label"] == "NPS 0-10"
    assert payload["variables"][2]["value_labels"]["price"] == "Pris"
    assert payload["rows"][1]["nps_score"] == 3


def test_xlsx_roundtrip_with_codebook(dataset):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(export_xlsx(dataset)))
    data = wb["data"]
    assert data.cell(1, 2).value == "nps_score"
    assert data.cell(2, 2).value == 10
    assert str(data.cell(5, 1).value).startswith("'="), "xlsx formula cells must be neutralized"
    codebook = wb["codebook"]
    labels = {codebook.cell(r, 1).value: codebook.cell(r, 2).value for r in range(2, 5)}
    assert labels["nps_score"] == "NPS 0-10"


def test_sav_roundtrip(dataset):
    blob = export_sav(dataset)
    back = import_sav(blob)
    names = [v.name for v in back.variables]
    assert names == ["respondent_key", "nps_score", "main_reason"]
    labels = {v.name: v.label for v in back.variables}
    assert labels["nps_score"] == "NPS 0-10"
    value_labels = {v.name: v.value_labels for v in back.variables}
    assert value_labels["nps_score"].get("10.0") == "Extremely likely"
    assert back.rows[0]["nps_score"] == 10
    assert back.rows[2]["nps_score"] is None
    assert back.rows[1]["main_reason"] == "service"
